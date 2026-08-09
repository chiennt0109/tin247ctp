import tempfile
from pathlib import Path
from unittest.mock import patch

from django.test import SimpleTestCase
from openpyxl import Workbook, load_workbook

from assessment.services.bank_importer import BankValidationError, WorkbookBankImporter


REQUIRED_HEADERS = {
    "FILES": [
        "FILE_ID", "FILE_NAME", "MIME_TYPE", "PARENT_FOLDER_ID", "FOLDER_PATH",
        "DRIVE_URL", "SOURCE_GROUP", "FILE_STATUS", "CHECKSUM", "CREATED_AT",
        "MODIFIED_AT", "INDEXED_AT", "NOTE",
    ],
    "CURRICULUM": ["CURRICULUM_ID", "GRADE", "SUBJECT", "PROGRAM_VERSION", "TOPIC_CODE", "TOPIC_NAME", "ORDER_NO", "STATUS", "NOTE"],
    "CURRICULUM_OUTCOMES": ["OUTCOME_ID", "CURRICULUM_ID", "OUTCOME_CODE", "OUTCOME_TEXT", "LEVEL", "STATUS", "NOTE"],
    "QUESTIONS": ["QUESTION_ID", "QUESTION_CODE", "QUESTION_TYPE", "COGNITIVE_LEVEL", "STEM_TEXT", "ANSWER_KEY", "EXPLANATION_ID", "STATUS", "VERSION", "LANGUAGE", "CREATED_AT", "UPDATED_AT", "NOTE", "DIFFICULTY", "COMPETENCY", "ESTIMATED_TIME_SEC", "USE_PURPOSE", "SHUFFLE_ALLOWED", "FAMILY_ID", "PROCESS_STATUS", "CLASSIFICATION_BASIS"],
    "OPTIONS": ["OPTION_ID", "QUESTION_ID", "OPTION_LABEL", "OPTION_TEXT", "IS_CORRECT", "ORDER_NO", "STATUS"],
    "STATEMENTS": ["STATEMENT_ID", "QUESTION_ID", "STATEMENT_LABEL", "STATEMENT_TEXT", "TRUTH_VALUE", "ORDER_NO", "STATUS", "COGNITIVE_LEVEL", "DIFFICULTY", "CLASSIFICATION_BASIS"],
    "QUESTION_CURRICULUM": ["QUESTION_CURRICULUM_ID", "QUESTION_ID", "CURRICULUM_ID", "OUTCOME_ID", "WEIGHT", "STATUS", "NOTE"],
    "QUESTION_SOURCES": ["QUESTION_SOURCE_ID", "QUESTION_ID", "FILE_ID", "SOURCE_PAGE", "SOURCE_SECTION", "SOURCE_REF", "LICENSE_NOTE", "STATUS"],
    "DUPLICATES": ["DUPLICATE_ID"], "POLICY_PROFILES": ["POLICY_PROFILE_ID"],
    "SCORE_RULES": ["SCORE_RULE_ID"], "BLUEPRINTS": ["BLUEPRINT_ID"],
    "BLUEPRINT_CELLS": ["BLUEPRINT_CELL_ID"], "BLUEPRINT_SLOTS": ["BLUEPRINT_SLOT_ID"],
    "QUY_UOC": ["CONFIG_KEY"],
}


class WorkbookFactory:
    @staticmethod
    def create(
        *, missing_answer=False, duplicate_question=False, duplicate_source=False,
        estimated_time=60, reorder_question_headers=False, source_group="", file_note="",
        checksum="",
    ):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in REQUIRED_HEADERS.items():
            sheet = workbook.create_sheet(name)
            actual_headers = list(headers)
            if name == "QUESTIONS" and reorder_question_headers:
                estimated_index = actual_headers.index("ESTIMATED_TIME_SEC")
                purpose_index = actual_headers.index("USE_PURPOSE")
                actual_headers[estimated_index], actual_headers[purpose_index] = (
                    actual_headers[purpose_index], actual_headers[estimated_index]
                )
            sheet.append(actual_headers)
        workbook["FILES"].append([
            "F1", "source.pdf", "application/pdf", "PARENT", "/source",
            "https://example.com/f", source_group, "PARSED", checksum, "", "", "", file_note,
        ])
        workbook["CURRICULUM"].append(["C1", 12, "Tin học", "GDPT2018", "A", "Topic", 1, "REVIEW", ""])
        workbook["CURRICULUM_OUTCOMES"].append(["O1", "C1", "YCCD_01", "Outcome", "BIET", "REVIEW", ""])
        question = dict(zip(REQUIRED_HEADERS["QUESTIONS"], [
            "Q1", "Q1", "MCQ_SINGLE", "BIET", "Stem", None if missing_answer else "A",
            "", "ACTIVE", 1, "vi", "", "", "", 1, "NLa", estimated_time, "PRACTICE",
            True, "FAM1", "READY_FOR_PRACTICE", "YCCD",
        ]))
        question_headers = [cell.value for cell in workbook["QUESTIONS"][1]]
        question_row = [question[header] for header in question_headers]
        workbook["QUESTIONS"].append(question_row)
        if duplicate_question:
            workbook["QUESTIONS"].append(question_row)
        for index, label in enumerate("ABCD", 1):
            workbook["OPTIONS"].append([f"OP{index}", "Q1", label, label, label == "A", index, "APPROVED"])
        workbook["QUESTION_CURRICULUM"].append(["QC1", "Q1", "C1", "O1", 1, "APPROVED", ""])
        workbook["QUESTION_SOURCES"].append(["QS1", "Q1", "F1", "1", "", "", "", "APPROVED"])
        if duplicate_source:
            workbook["QUESTION_SOURCES"].append(
                ["QS2", "Q1", "F1", "2", "updated-section", "updated-ref", "updated-license", "APPROVED"]
            )
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        return Path(handle.name)


class WorkbookBankImporterTests(SimpleTestCase):
    def test_parses_valid_question_and_stable_hash(self):
        path = WorkbookFactory.create()
        self.addCleanup(path.unlink)
        first = WorkbookBankImporter().parse(path)
        second = WorkbookBankImporter().parse(path)
        self.assertFalse(first.errors)
        self.assertEqual(first.questions[0]["content_hash"], second.questions[0]["content_hash"])

    def test_rejects_missing_answer(self):
        path = WorkbookFactory.create(missing_answer=True)
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertIn("MISSING_ANSWER", parsed.errors[0]["issues"])

    def test_imports_periodic_essay_without_options_and_maps_physical_review_status(self):
        path = WorkbookFactory.create()
        workbook = load_workbook(path)
        headers = [cell.value for cell in workbook["QUESTIONS"][1]]
        row = dict(zip(headers, next(workbook["QUESTIONS"].iter_rows(min_row=2, values_only=True))))
        row.update({
            "QUESTION_ID": "Q_ESSAY", "QUESTION_CODE": "Q_ESSAY",
            "QUESTION_TYPE": "ESSAY", "ANSWER_KEY": "Rubric: đủ 3 ý",
            "STATUS": "ACTIVE", "PROCESS_STATUS": None,
            "USE_PURPOSE": "PERIODIC", "SHUFFLE_ALLOWED": False,
            "FAMILY_ID": "ESSAY_FAMILY",
        })
        workbook["QUESTIONS"].append([row.get(header) for header in headers])
        workbook["QUESTION_CURRICULUM"].append(
            ["QC_ESSAY", "Q_ESSAY", "C1", "O1", 1, "APPROVED", ""]
        )
        workbook["QUESTION_SOURCES"].append(
            ["QS_ESSAY", "Q_ESSAY", "F1", "1", "", "", "", "APPROVED"]
        )
        workbook.save(path)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)

        essay = next(question for question in parsed.questions if question["question_id"] == "Q_ESSAY")
        self.assertEqual(essay["options"], [])
        self.assertEqual(essay["process_status"], "READY_FOR_PERIODIC")
        self.assertEqual(essay["row"]["STATUS"], "ACTIVE")
        self.assertIsNone(essay["row"]["__source_process_status__"])
        self.assertTrue(essay["row"]["__process_status_derived__"])
        warning = next(item for item in parsed.warnings if item["question_id"] == "Q_ESSAY")
        self.assertEqual(warning["derived_value"], "READY_FOR_PERIODIC")
        self.assertFalse(any(error.get("question_id") == "Q_ESSAY" for error in parsed.errors))

    def test_blank_process_status_is_not_derived_for_draft_periodic_question(self):
        path = WorkbookFactory.create()
        workbook = load_workbook(path)
        headers = [cell.value for cell in workbook["QUESTIONS"][1]]
        workbook["QUESTIONS"].cell(2, headers.index("STATUS") + 1, "DRAFT")
        workbook["QUESTIONS"].cell(2, headers.index("USE_PURPOSE") + 1, "PERIODIC")
        workbook["QUESTIONS"].cell(2, headers.index("PROCESS_STATUS") + 1).value = None
        workbook.save(path)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)

        error = next(item for item in parsed.errors if item.get("question_id") == "Q1")
        self.assertIn("INVALID_PROCESS_STATUS", error["issues"])
        self.assertFalse(parsed.warnings)

    def test_rejects_duplicate_question_key(self):
        path = WorkbookFactory.create(duplicate_question=True)
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertTrue(any(error["code"] == "DUPLICATE_KEY" for error in parsed.errors))

    def test_single_outcome_physical_row_is_not_a_duplicate(self):
        path = WorkbookFactory.create()
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)

        self.assertEqual(parsed.key_rows["CURRICULUM_OUTCOMES"]["O1"], [2])
        self.assertFalse(any(
            error["code"] == "DUPLICATE_KEY" and error["sheet"] == "CURRICULUM_OUTCOMES"
            for error in parsed.errors
        ))

    def test_duplicate_outcome_reports_distinct_physical_rows_and_values(self):
        path = WorkbookFactory.create()
        workbook = load_workbook(path)
        workbook["CURRICULUM_OUTCOMES"].append(
            ["Ｏ1 ", "C1", "YCCD_02", "Duplicate", "BIET", "REVIEW", ""],
        )
        workbook.save(path)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)
        error = next(
            error for error in parsed.errors
            if error["code"] == "DUPLICATE_KEY" and error["sheet"] == "CURRICULUM_OUTCOMES"
        )

        self.assertEqual(error["key"], "O1")
        self.assertEqual(error["count"], 2)
        self.assertEqual(error["row_numbers"], [2, 3])
        self.assertEqual(error["raw_values"], ["O1", "Ｏ1 "])
        self.assertEqual(error["normalized_values"], ["O1", "O1"])

    def test_same_physical_outcome_row_appended_twice_is_not_duplicate(self):
        records = [{
            "OUTCOME_ID": "O1", "__row__": 133,
            "__key_raw__": "O1", "__key_normalized__": "O1",
        }]
        errors = []

        indexes = WorkbookBankImporter._validate_unique_keys(
            {"CURRICULUM_OUTCOMES": records + records}, errors,
        )

        self.assertEqual(indexes["CURRICULUM_OUTCOMES"]["O1"], [133])
        self.assertEqual(errors, [])

    def test_question_and_blueprint_references_do_not_duplicate_outcome(self):
        path = WorkbookFactory.create()
        workbook = load_workbook(path)
        sheet = workbook["BLUEPRINT_CELLS"]
        sheet.cell(row=1, column=2, value="OUTCOME_ID")
        sheet.append(["BC1", "O1"])
        workbook.save(path)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)

        self.assertEqual(parsed.key_rows["CURRICULUM_OUTCOMES"]["O1"], [2])
        self.assertFalse(any(
            error["code"] == "DUPLICATE_KEY" and error.get("key") == "O1"
            for error in parsed.errors
        ))

    def test_curriculum_outcomes_worksheet_is_parsed_once(self):
        path = WorkbookFactory.create()
        self.addCleanup(path.unlink)
        from assessment.services import bank_importer

        with patch.object(bank_importer, "_sheet_rows", wraps=bank_importer._sheet_rows) as reader:
            WorkbookBankImporter().parse(path)

        outcome_calls = [call for call in reader.call_args_list if call.args[1] == "CURRICULUM_OUTCOMES"]
        self.assertEqual(len(outcome_calls), 1)

    def test_rejects_missing_required_sheet(self):
        path = WorkbookFactory.create()
        workbook = load_workbook(path)
        del workbook["QUESTIONS"]
        workbook.save(path)
        self.addCleanup(path.unlink)
        with self.assertRaises(BankValidationError):
            WorkbookBankImporter().parse(path)

    def test_deduplicates_question_sources_by_question_and_file(self):
        path = WorkbookFactory.create(duplicate_source=True)
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertFalse(parsed.errors)
        self.assertEqual(len(parsed.questions[0]["sources"]), 1)
        self.assertEqual(parsed.questions[0]["sources"][0]["SOURCE_PAGE"], "2")

    def test_estimated_time_accepts_integer(self):
        path = WorkbookFactory.create(estimated_time=75)
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertFalse(parsed.errors)
        self.assertEqual(parsed.questions[0]["estimated_time_seconds"], 75)

    def test_dry_run_reports_remaining_char_field_overflow(self):
        path = WorkbookFactory.create(checksum="x" * 129)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)

        error = next(error for error in parsed.errors if error["code"] == "FIELD_TOO_LONG")
        self.assertEqual(error["sheet"], "FILES")
        self.assertEqual(error["row"], 2)
        self.assertEqual(error["column"], "CHECKSUM")
        self.assertEqual(error["field"], "checksum")
        self.assertEqual(error["model_field"], "checksum")
        self.assertEqual(error["length"], 129)
        self.assertEqual(error["max_length"], 128)
        self.assertEqual(error["cell"], "I2")
        self.assertEqual(error["header_index"], 9)
        self.assertEqual(error["raw_value_preview"], "x" * 120)

    def test_blank_checksum_and_long_note_keep_their_physical_columns(self):
        note = "N" * 1000
        path = WorkbookFactory.create(file_note=note, checksum="")
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)
        source = parsed.rows["FILES"][0]

        self.assertFalse(parsed.errors)
        self.assertEqual(source["NOTE"], note)
        self.assertIsNone(source["CHECKSUM"])
        self.assertEqual(source["FILE_STATUS"], "PARSED")

    def test_multiple_middle_blanks_do_not_shift_file_columns(self):
        path = WorkbookFactory.create(source_group="group", file_note="note")
        workbook = load_workbook(path)
        sheet = workbook["FILES"]
        sheet["C2"] = None  # MIME_TYPE
        sheet["D2"] = None  # PARENT_FOLDER_ID
        sheet["E2"] = None  # FOLDER_PATH
        sheet["F2"] = None  # DRIVE_URL
        sheet["I2"] = None  # CHECKSUM
        workbook.save(path)
        self.addCleanup(path.unlink)

        source = WorkbookBankImporter().parse(path).rows["FILES"][0]

        self.assertIsNone(source["MIME_TYPE"])
        self.assertIsNone(source["DRIVE_URL"])
        self.assertIsNone(source["FOLDER_PATH"])
        self.assertEqual(source["SOURCE_GROUP"], "group")
        self.assertEqual(source["NOTE"], "note")
        self.assertIsNone(source["CHECKSUM"])
        self.assertEqual(source["FILE_STATUS"], "PARSED")

    def test_invalid_checksum_format_is_fatal(self):
        path = WorkbookFactory.create(checksum="not-a-sha256")
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)
        error = next(error for error in parsed.errors if error["code"] == "INVALID_CHECKSUM")

        self.assertEqual(error["cell"], "I2")
        self.assertEqual(error["header_index"], 9)
        self.assertEqual(error["model_field"], "checksum")

    def test_tagged_legacy_checksum_payload_is_preserved_as_note(self):
        payload = "COMPLETED_SOURCE_DETAILS=" + "THANH_HOA," * 60
        path = WorkbookFactory.create(checksum=payload)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)
        source = parsed.rows["FILES"][0]

        self.assertFalse(parsed.errors)
        self.assertIsNone(source["CHECKSUM"])
        self.assertEqual(source["NOTE"], payload)

    def test_estimated_time_blank_becomes_none(self):
        path = WorkbookFactory.create(estimated_time="")
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertFalse(parsed.errors)
        self.assertIsNone(parsed.questions[0]["estimated_time_seconds"])

    def test_estimated_time_accepts_numeric_string(self):
        path = WorkbookFactory.create(estimated_time="90")
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertFalse(parsed.errors)
        self.assertEqual(parsed.questions[0]["estimated_time_seconds"], 90)

    def test_estimated_time_rejects_purpose_value_during_dry_run(self):
        path = WorkbookFactory.create(estimated_time="PRACTICE")
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertTrue(parsed.has_fatal_errors)
        self.assertTrue(any(
            error.get("code") == "INVALID_FIELD_TYPE"
            and error.get("field") == "ESTIMATED_TIME_SEC"
            and error.get("value") == "PRACTICE"
            for error in parsed.errors
        ))

    def test_question_columns_are_mapped_by_normalized_header(self):
        path = WorkbookFactory.create(estimated_time="120", reorder_question_headers=True)
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertFalse(parsed.errors)
        self.assertEqual(parsed.questions[0]["estimated_time_seconds"], 120)
        self.assertEqual(parsed.questions[0]["use_purpose"], "PRACTICE")
