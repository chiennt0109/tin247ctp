from io import StringIO

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from assessment.models import (
    BankQuestion, BankQuestionRevision, BankSourceFile, CurriculumNode,
    QuestionAsset, QuestionSyncLog,
)
from assessment.services.bank_importer import WorkbookBankImporter
from assessment.services.bank_sync import BankSyncService
from assessment.tests.test_bank_importer import WorkbookFactory


class BankSyncServiceTests(TestCase):
    def setUp(self):
        self.path = WorkbookFactory.create()
        self.addCleanup(self.path.unlink)

    def test_preview_does_not_write(self):
        parsed = WorkbookBankImporter().parse(self.path)
        report = BankSyncService().preview(parsed)
        self.assertEqual(report["new"], 1)
        self.assertEqual(BankQuestion.objects.count(), 0)
        self.assertEqual(QuestionSyncLog.objects.count(), 0)

    def test_apply_creates_projection_and_unchanged_sync_does_not_add_revision(self):
        parsed = WorkbookBankImporter().parse(self.path)
        BankSyncService().apply(parsed)
        question = BankQuestion.objects.get(source_question_id="Q1")
        self.assertTrue(question.is_available)
        self.assertEqual(question.current_revision.protected_answer, {"answer_key": "A"})
        BankSyncService().apply(WorkbookBankImporter().parse(self.path))
        self.assertEqual(BankQuestionRevision.objects.filter(question=question).count(), 1)

    def test_existing_database_rows_are_unchanged_not_duplicate_source_keys(self):
        BankSyncService().apply(WorkbookBankImporter().parse(self.path))

        parsed = WorkbookBankImporter().parse(self.path)
        report = BankSyncService().preview(parsed)

        self.assertEqual(report["unchanged"], 1)
        self.assertEqual(report.get("new", 0), 0)
        self.assertFalse(any(error["code"] == "DUPLICATE_KEY" for error in report["errors"]))

    def test_thirteen_new_questions_are_created_exactly_once(self):
        workbook = load_workbook(self.path)
        question_sheet = workbook["QUESTIONS"]
        question_headers = [cell.value for cell in question_sheet[1]]
        original_question = dict(zip(question_headers, next(question_sheet.iter_rows(min_row=2, values_only=True))))
        for number in range(2, 14):
            question_id = f"Q{number}"
            question = dict(original_question)
            question.update({
                "QUESTION_ID": question_id, "QUESTION_CODE": question_id,
                "FAMILY_ID": f"FAM{number}",
            })
            question_sheet.append([question.get(header) for header in question_headers])
            for index, label in enumerate("ABCD", 1):
                workbook["OPTIONS"].append([
                    f"{question_id}-OP{index}", question_id, label, label,
                    label == "A", index, "APPROVED",
                ])
            workbook["QUESTION_CURRICULUM"].append([
                f"{question_id}-QC", question_id, "C1", "O1", 1, "APPROVED", "",
            ])
            workbook["QUESTION_SOURCES"].append([
                f"{question_id}-QS", question_id, "F1", "1", "", "", "", "APPROVED",
            ])
        workbook.save(self.path)

        first = WorkbookBankImporter().parse(self.path)
        self.assertEqual(len(first.questions), 13)
        BankSyncService().apply(first)
        second = WorkbookBankImporter().parse(self.path)
        second_report = BankSyncService().preview(second)
        BankSyncService().apply(second)

        self.assertEqual(BankQuestion.objects.count(), 13)
        self.assertEqual(BankQuestionRevision.objects.count(), 13)
        self.assertEqual(second_report["unchanged"], 13)
        self.assertEqual(second_report.get("new", 0), 0)

    def test_apply_uses_the_integer_normalized_by_dry_run_pipeline(self):
        numeric_string_path = WorkbookFactory.create(estimated_time="135")
        self.addCleanup(numeric_string_path.unlink)
        parsed = WorkbookBankImporter().parse(numeric_string_path)
        self.assertFalse(parsed.errors)
        self.assertEqual(parsed.questions[0]["estimated_time_seconds"], 135)
        BankSyncService().apply(parsed)
        self.assertEqual(BankQuestion.objects.get().estimated_time_seconds, 135)

    def test_long_file_source_group_and_note_are_preserved(self):
        source_group = "Nhóm nguồn " + "g" * 450
        note = "Ghi chú " + "n" * 1200
        path = WorkbookFactory.create(source_group=source_group, file_note=note)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)
        self.assertFalse(parsed.errors)
        BankSyncService().apply(parsed)

        source = BankSourceFile.objects.get(source_id="F1")
        self.assertEqual(source.source_group, source_group)
        self.assertEqual(source.note, note)

    def test_valid_checksum_is_saved_unchanged(self):
        checksum = "SHA256:" + "a1" * 32
        path = WorkbookFactory.create(checksum=checksum)
        self.addCleanup(path.unlink)

        parsed = WorkbookBankImporter().parse(path)
        self.assertFalse(parsed.errors)
        BankSyncService().apply(parsed)

        self.assertEqual(BankSourceFile.objects.get(source_id="F1").checksum, checksum)

    def test_validation_error_prevents_all_apply_writes(self):
        path = WorkbookFactory.create(checksum="invalid")
        self.addCleanup(path.unlink)
        parsed = WorkbookBankImporter().parse(path)
        self.assertTrue(parsed.has_fatal_errors)

        with self.assertRaises(ValueError):
            BankSyncService().apply(parsed)

        self.assertFalse(BankSourceFile.objects.exists())
        self.assertFalse(CurriculumNode.objects.exists())
        self.assertFalse(BankQuestion.objects.exists())

    def test_mid_sync_failure_rolls_back_files_curriculum_and_questions(self):
        parsed = WorkbookBankImporter().parse(self.path)

        from unittest.mock import patch
        with patch.object(BankSyncService, "_sync_question", side_effect=RuntimeError("mid-sync")):
            with self.assertRaisesMessage(ValueError, "mid-sync"):
                BankSyncService().apply(parsed)

        self.assertFalse(BankSourceFile.objects.exists())
        self.assertFalse(CurriculumNode.objects.exists())
        self.assertFalse(BankQuestion.objects.exists())
        self.assertFalse(QuestionSyncLog.objects.exists())

    def test_changed_content_creates_revision_and_preserves_old_revision(self):
        BankSyncService().apply(WorkbookBankImporter().parse(self.path))
        workbook = load_workbook(self.path)
        workbook["QUESTIONS"]["E2"] = "Updated stem"
        workbook.save(self.path)
        BankSyncService().apply(WorkbookBankImporter().parse(self.path))
        question = BankQuestion.objects.get(source_question_id="Q1")
        self.assertEqual(question.revisions.count(), 2)
        self.assertEqual(question.current_revision.stem_text, "Updated stem")

    def test_fatal_validation_error_rolls_back_everything(self):
        workbook = load_workbook(self.path)
        workbook["QUESTIONS"]["F2"] = None
        workbook.save(self.path)
        parsed = WorkbookBankImporter().parse(self.path)
        with self.assertRaises(ValueError):
            BankSyncService().apply(parsed)
        self.assertEqual(BankQuestion.objects.count(), 0)
        self.assertEqual(QuestionSyncLog.objects.count(), 0)

    @override_settings(QUESTION_BANK_SYNC_ENABLED=True)
    def test_command_apply_twice_deduplicates_source_and_updates_asset(self):
        duplicate_path = WorkbookFactory.create(duplicate_source=True)
        self.addCleanup(duplicate_path.unlink)

        call_command("sync_exam_bank", "--source", str(duplicate_path), "--apply", verbosity=0)
        asset = QuestionAsset.objects.get()
        original_pk = asset.pk
        self.assertEqual(asset.source_page, "2")
        self.assertEqual(QuestionAsset.objects.count(), 1)

        # A subsequent master update for the same (question, source_file) must
        # update the existing relation rather than attempting another INSERT.
        workbook = load_workbook(duplicate_path)
        workbook["QUESTION_SOURCES"]["D3"] = "3"
        workbook["QUESTION_SOURCES"]["E3"] = "newest-section"
        workbook.save(duplicate_path)
        call_command("sync_exam_bank", "--source", str(duplicate_path), "--apply", verbosity=0)

        asset.refresh_from_db()
        self.assertEqual(asset.pk, original_pk)
        self.assertEqual(asset.source_page, "3")
        self.assertEqual(asset.source_section, "newest-section")
        self.assertEqual(QuestionAsset.objects.count(), 1)

    @override_settings(QUESTION_BANK_SYNC_ENABLED=True)
    def test_apply_validation_failure_reports_failed_and_writes_nothing(self):
        invalid_path = WorkbookFactory.create(checksum="x" * 129)
        self.addCleanup(invalid_path.unlink)
        output = StringIO()

        with self.assertRaises(CommandError):
            call_command(
                "sync_exam_bank", "--source", str(invalid_path), "--apply",
                stdout=output, verbosity=0,
            )

        self.assertIn('"mode": "APPLY_VALIDATION"', output.getvalue())
        self.assertIn('"mode": "APPLY_FAILED"', output.getvalue())
        self.assertNotIn('"mode": "APPLY_SUCCESS"', output.getvalue())
        self.assertFalse(BankSourceFile.objects.exists())
        self.assertFalse(QuestionSyncLog.objects.exists())
