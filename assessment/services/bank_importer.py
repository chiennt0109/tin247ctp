import hashlib
import json
import math
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


QUESTION_TYPES = {"MCQ_SINGLE", "TRUE_FALSE_GROUP", "SHORT_ANSWER", "ESSAY", "PRACTICAL"}
COGNITIVE_LEVELS = {"BIET", "HIEU", "VANDUNG"}
COMPETENCIES = {"NLa", "NLb", "NLc", "NLd", "NLe"}
PROCESS_STATUSES = {
    "RAW", "NORMALIZED", "CURRICULUM_MAPPED", "CLASSIFIED", "DUPLICATE_CHECKED",
    "ANSWER_CHECKED", "CONTENT_REVIEWED", "READY_FOR_PRACTICE", "READY_FOR_PERIODIC",
    "READY_FOR_GRADUATION", "NEEDS_REVIEW", "OUTDATED", "RETIRED",
}
STATUS_VALUES = {"DRAFT", "PENDING", "REVIEW", "APPROVED", "ACTIVE", "INACTIVE", "REJECTED", "ARCHIVED"}
USE_PURPOSES = {"PRACTICE", "PERIODIC", "GRADUATION", "NONE", "REVIEW_ONLY"}

# Types used by either validation or persistence. Normalization happens once in
# the importer, so dry-run and apply consume the exact same typed values.
INTEGER_FIELDS = {
    "CURRICULUM": ("GRADE", "ORDER_NO"),
    "QUESTIONS": ("VERSION", "DIFFICULTY", "ESTIMATED_TIME_SEC"),
    "OPTIONS": ("ORDER_NO",),
    "STATEMENTS": ("ORDER_NO", "DIFFICULTY"),
}
DECIMAL_FIELDS = {
    "QUESTION_CURRICULUM": ("WEIGHT",),
}
BOOLEAN_FIELDS = {
    "QUESTIONS": ("SHUFFLE_ALLOWED",),
    "OPTIONS": ("IS_CORRECT",),
    "STATEMENTS": ("TRUTH_VALUE",),
}
DATE_FIELDS = {
    "QUESTIONS": ("CREATED_AT", "UPDATED_AT"),
}
REQUIRED_SHEETS = {
    "FILES", "CURRICULUM", "CURRICULUM_OUTCOMES", "QUESTIONS", "OPTIONS", "STATEMENTS",
    "QUESTION_CURRICULUM", "QUESTION_SOURCES", "DUPLICATES", "POLICY_PROFILES",
    "SCORE_RULES", "BLUEPRINTS", "BLUEPRINT_CELLS", "BLUEPRINT_SLOTS", "QUY_UOC",
}

MODEL_FIELD_MAPPINGS = {
    "FILES": {
        "FILE_ID": "source_id", "FILE_NAME": "name", "MIME_TYPE": "mime_type",
        "DRIVE_URL": "drive_url", "FOLDER_PATH": "folder_path",
        "SOURCE_GROUP": "source_group", "NOTE": "note", "CHECKSUM": "checksum",
        "FILE_STATUS": "source_status",
    },
}
FILES_NOTE_MARKERS = ("COMPLETED_SOURCE_DETAILS=",)


class BankValidationError(ValueError):
    pass


def _canonical(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _boolean(value):
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().upper()
    if normalized in {"TRUE", "1", "YES"}:
        return True
    if normalized in {"FALSE", "0", "NO"}:
        return False
    raise BankValidationError(f"Invalid boolean value: {value!r}")


def _optional_integer(value):
    """Parse an optional integer without accepting lossy or boolean values."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        raise BankValidationError(f"Expected integer, got boolean {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return int(value)
        raise BankValidationError(f"Expected integer, got {value!r}")
    if isinstance(value, str):
        text = value.strip()
        try:
            decimal = Decimal(text)
        except InvalidOperation as exc:
            raise BankValidationError(f"Expected integer, got {value!r}") from exc
        if decimal.is_finite() and decimal == decimal.to_integral_value():
            return int(decimal)
    raise BankValidationError(f"Expected integer, got {value!r}")


def _optional_decimal(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        raise BankValidationError(f"Expected decimal, got boolean {value!r}")
    try:
        decimal = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise BankValidationError(f"Expected decimal, got {value!r}") from exc
    if not decimal.is_finite():
        raise BankValidationError(f"Expected finite decimal, got {value!r}")
    return decimal


def _optional_datetime(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError as exc:
            raise BankValidationError(f"Expected ISO date/datetime, got {value!r}") from exc
    raise BankValidationError(f"Expected date/datetime, got {value!r}")


def _normalize_header(value):
    if value is None:
        return ""
    normalized = unicodedata.normalize("NFKC", str(value)).lstrip("\ufeff").strip().upper()
    return "_".join(normalized.split())


def _normalize_source_key(value):
    """Normalize a physical sheet key exactly once, without changing its case."""
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def _sheet_rows(workbook, name):
    sheet = workbook[name]
    iterator = sheet.iter_rows(values_only=True)
    headers = [_normalize_header(value) for value in next(iterator)]
    populated_headers = [header for header in headers if header]
    if len(populated_headers) != len(set(populated_headers)):
        raise BankValidationError(f"Sheet {name} contains duplicate normalized headers")
    rows = []
    for row_number, values in enumerate(iterator, start=2):
        # Positional mapping is deliberate: never compact/filter None values,
        # otherwise a blank CHECKSUM can shift NOTE into the checksum column.
        row_values = list(values)
        record = {
            headers[index]: _canonical(row_values[index] if index < len(row_values) else None)
            for index in range(len(headers))
            if headers[index]
        }
        record["__cells__"] = {
            headers[index]: f"{get_column_letter(index + 1)}{row_number}"
            for index in range(len(headers)) if headers[index]
        }
        record["__header_indexes__"] = {
            headers[index]: index + 1 for index in range(len(headers)) if headers[index]
        }
        # Master rule: the first column is the unique key. Formatting/formula-only rows are ignored.
        raw_key = record.get(headers[0])
        normalized_key = _normalize_source_key(raw_key)
        if not normalized_key:
            continue
        record[headers[0]] = normalized_key
        record["__row__"] = row_number
        record["__key_raw__"] = str(raw_key)
        record["__key_normalized__"] = normalized_key
        rows.append(record)
    return rows


@dataclass
class ParsedBank:
    source_path: str
    source_sha256: str
    rows: dict
    questions: list
    errors: list
    warnings: list
    key_rows: dict

    @property
    def has_fatal_errors(self):
        return bool(self.errors)


class WorkbookBankImporter:
    def parse(self, source_path):
        source_path = str(Path(source_path))
        payload = Path(source_path).read_bytes()
        raw = load_workbook(source_path, read_only=True, data_only=False)
        cached = load_workbook(source_path, read_only=True, data_only=True)
        missing = sorted(REQUIRED_SHEETS - set(cached.sheetnames))
        if missing:
            raise BankValidationError(f"Missing required sheets: {', '.join(missing)}")

        rows = {name: _sheet_rows(cached, name) for name in cached.sheetnames}
        raw_rows = {name: _sheet_rows(raw, name) for name in ("QUESTIONS",)}
        errors, warnings = [], []
        key_rows = self._validate_unique_keys(rows, errors)
        self._normalize_and_validate_types(rows, errors)
        self._normalize_file_metadata(rows)
        self._validate_model_field_lengths(rows, errors)
        self._validate_file_checksums(rows, errors)
        questions = self._build_questions(rows, raw_rows, errors, warnings)
        return ParsedBank(
            source_path=source_path,
            source_sha256=hashlib.sha256(payload).hexdigest(),
            rows=rows,
            questions=questions,
            errors=errors,
            warnings=warnings,
            key_rows=key_rows,
        )

    @staticmethod
    def _validate_unique_keys(rows, errors):
        indexes = {}
        for sheet, records in rows.items():
            if not records:
                indexes[sheet] = {}
                continue
            key = next(field for field in records[0] if not field.startswith("__"))
            occurrences = defaultdict(dict)
            for record in records:
                normalized = record.get("__key_normalized__", _normalize_source_key(record.get(key)))
                row_number = record.get("__row__")
                # Row number is the physical identity. If a downstream list is
                # accidentally concatenated with itself, the same physical row
                # must not become a duplicate.
                occurrences[normalized].setdefault(row_number, {
                    "raw": record.get("__key_raw__", record.get(key)),
                    "normalized": normalized,
                })
            indexes[sheet] = {
                normalized: sorted(by_row)
                for normalized, by_row in occurrences.items()
            }
            for value, by_row in occurrences.items():
                row_numbers = sorted(by_row)
                if len(row_numbers) > 1:
                    errors.append({
                        "code": "DUPLICATE_KEY",
                        "sheet": sheet,
                        "key": value,
                        "count": len(row_numbers),
                        "row_numbers": row_numbers,
                        "raw_values": [by_row[row]["raw"] for row in row_numbers],
                        "normalized_values": [by_row[row]["normalized"] for row in row_numbers],
                    })
        return indexes

    @staticmethod
    def _validate_model_field_lengths(rows, errors):
        # Import lazily so the parser module remains usable by workbook-only
        # tooling while Django is being initialized.
        from assessment.models import BankSourceFile

        models_by_sheet = {"FILES": BankSourceFile}
        for sheet, mapping in MODEL_FIELD_MAPPINGS.items():
            model = models_by_sheet[sheet]
            for row in rows.get(sheet, ()):
                for column, field_name in mapping.items():
                    value = row.get(column)
                    if value in (None, ""):
                        continue
                    field = model._meta.get_field(field_name)
                    max_length = getattr(field, "max_length", None)
                    if max_length is not None and len(str(value)) > max_length:
                        raw = str(value)
                        errors.append({
                            "code": "FIELD_TOO_LONG", "sheet": sheet,
                            "row": row.get("__row__"), "column": column,
                            "field": field_name, "length": len(str(value)),
                            "max_length": max_length,
                            "cell": row.get("__cells__", {}).get(column),
                            "raw_value_preview": raw[:120],
                            "header_index": row.get("__header_indexes__", {}).get(column),
                            "model_field": field_name,
                        })

    @staticmethod
    def _normalize_file_metadata(rows):
        """Handle tagged metadata stored in the legacy CHECKSUM cell.

        The canonical master has historical FILES rows where a clearly tagged
        source-detail payload was written to CHECKSUM (for example
        ``COMPLETED_SOURCE_DETAILS=...``) while NOTE is empty. This is not a
        checksum and must be preserved losslessly as note metadata. Arbitrary
        invalid checksum values are never moved and still fail validation.
        """
        for row in rows.get("FILES", ()):
            checksum = row.get("CHECKSUM")
            if checksum in (None, ""):
                continue
            raw = str(checksum)
            if not raw.startswith(FILES_NOTE_MARKERS):
                continue
            note = str(row.get("NOTE") or "")
            row["NOTE"] = f"{note}\n{raw}".strip() if note else raw
            row["CHECKSUM"] = None

    @staticmethod
    def _validate_file_checksums(rows, errors):
        checksum_pattern = re.compile(r"^(?:[0-9a-fA-F]{64}|SHA256:[0-9a-fA-F]{64})$")
        for row in rows.get("FILES", ()):
            value = row.get("CHECKSUM")
            if value in (None, ""):
                continue
            raw = str(value)
            if not checksum_pattern.fullmatch(raw):
                errors.append({
                    "code": "INVALID_CHECKSUM", "sheet": "FILES",
                    "row": row.get("__row__"), "column": "CHECKSUM",
                    "cell": row.get("__cells__", {}).get("CHECKSUM"),
                    "raw_value_preview": raw[:120],
                    "header_index": row.get("__header_indexes__", {}).get("CHECKSUM"),
                    "model_field": "checksum",
                    "allowed": ["blank", "64 hex characters", "SHA256: + 64 hex characters"],
                })

    @staticmethod
    def _normalize_and_validate_types(rows, errors):
        specs = (
            (INTEGER_FIELDS, _optional_integer, "integer"),
            (DECIMAL_FIELDS, _optional_decimal, "decimal"),
            (BOOLEAN_FIELDS, _boolean, "boolean"),
            (DATE_FIELDS, _optional_datetime, "date/datetime"),
        )
        for sheet_fields, converter, expected in specs:
            for sheet, fields in sheet_fields.items():
                for row in rows.get(sheet, ()):
                    for field in fields:
                        if field not in row:
                            continue
                        original = row.get(field)
                        try:
                            row[field] = converter(original)
                        except BankValidationError:
                            row[field] = None
                            row.setdefault("__invalid_fields__", []).append(field)
                            error = {
                                "code": "INVALID_FIELD_TYPE", "sheet": sheet,
                                "row": row.get("__row__"), "field": field,
                                "value": str(original), "expected": expected,
                            }
                            if sheet == "QUESTIONS":
                                error["question_id"] = str(row.get("QUESTION_ID"))
                            errors.append(error)

        enum_specs = {
            "QUESTIONS": {
                "QUESTION_TYPE": QUESTION_TYPES, "COGNITIVE_LEVEL": COGNITIVE_LEVELS,
                "PROCESS_STATUS": PROCESS_STATUSES, "USE_PURPOSE": USE_PURPOSES,
                "STATUS": STATUS_VALUES, "COMPETENCY": COMPETENCIES,
            },
            "OPTIONS": {"STATUS": STATUS_VALUES},
            "STATEMENTS": {"COGNITIVE_LEVEL": COGNITIVE_LEVELS, "STATUS": STATUS_VALUES},
            "CURRICULUM_OUTCOMES": {"LEVEL": COGNITIVE_LEVELS},
        }
        for sheet, fields in enum_specs.items():
            for row in rows.get(sheet, ()):
                for field, allowed in fields.items():
                    value = row.get(field)
                    if value not in (None, "") and str(value) not in allowed:
                        errors.append({
                            "code": "INVALID_ENUM", "sheet": sheet, "row": row.get("__row__"),
                            "field": field, "value": str(value), "allowed": sorted(allowed),
                        })

    def _build_questions(self, rows, raw_rows, errors, warnings):
        curriculum_ids = {str(row.get("CURRICULUM_ID")) for row in rows["CURRICULUM"]}
        outcomes = {str(row.get("OUTCOME_ID")): str(row.get("CURRICULUM_ID")) for row in rows["CURRICULUM_OUTCOMES"]}
        file_ids = {str(row.get("FILE_ID")) for row in rows["FILES"]}
        options = defaultdict(list)
        for row in rows["OPTIONS"]:
            options[str(row.get("QUESTION_ID"))].append(row)
        statements = defaultdict(list)
        for row in rows["STATEMENTS"]:
            statements[str(row.get("QUESTION_ID"))].append(row)
        mappings = {str(row.get("QUESTION_ID")): row for row in rows["QUESTION_CURRICULUM"]}
        sources = defaultdict(list)
        for row in rows["QUESTION_SOURCES"]:
            sources[str(row.get("QUESTION_ID"))].append(row)
        # QUESTION_SOURCES has its own row key, but the operational relation is
        # unique per (question, source file). Normalize duplicate source rows
        # before handing them to the persistence service. The last row wins so
        # corrected metadata from a later master row is applied deterministically.
        for question_id, question_sources in tuple(sources.items()):
            deduplicated = {}
            for source in question_sources:
                deduplicated[str(source.get("FILE_ID"))] = source
            sources[question_id] = list(deduplicated.values())
        raw_by_id = {str(row.get("QUESTION_ID")): row for row in raw_rows["QUESTIONS"]}
        result = []
        for row in rows["QUESTIONS"]:
            qid = str(row.get("QUESTION_ID"))
            # Detailed type errors were already emitted by the shared typed-row
            # pipeline. Do not count an invalid row as a valid question.
            if row.get("__invalid_fields__"):
                continue
            qerrors = []
            qtype = str(row.get("QUESTION_TYPE") or "")
            if qtype not in QUESTION_TYPES:
                qerrors.append("INVALID_QUESTION_TYPE")
            if str(row.get("COGNITIVE_LEVEL") or "") not in COGNITIVE_LEVELS:
                qerrors.append("INVALID_COGNITIVE_LEVEL")
            if not str(row.get("STEM_TEXT") or "").strip():
                qerrors.append("MISSING_STEM")
            if row.get("ANSWER_KEY") in (None, ""):
                qerrors.append("MISSING_ANSWER")
            if str(row.get("PROCESS_STATUS") or "") not in PROCESS_STATUSES:
                qerrors.append("INVALID_PROCESS_STATUS")
            difficulty = row.get("DIFFICULTY")
            if difficulty not in range(1, 6):
                qerrors.append("INVALID_DIFFICULTY")
            shuffle_allowed = row.get("SHUFFLE_ALLOWED")
            if not isinstance(shuffle_allowed, bool):
                qerrors.append("INVALID_SHUFFLE_ALLOWED")
                shuffle_allowed = False
            if qtype == "MCQ_SINGLE":
                qoptions = options[qid]
                if len(qoptions) != 4 or sum(_boolean(item.get("IS_CORRECT")) for item in qoptions) != 1:
                    qerrors.append("INVALID_OPTIONS")
            if qtype == "TRUE_FALSE_GROUP" and len(statements[qid]) != 4:
                qerrors.append("INVALID_STATEMENTS")
            if qid not in mappings:
                qerrors.append("MISSING_CURRICULUM_LINK")
            else:
                mapping = mappings[qid]
                curriculum_id = str(mapping.get("CURRICULUM_ID"))
                outcome_id = str(mapping.get("OUTCOME_ID"))
                if curriculum_id not in curriculum_ids:
                    qerrors.append("UNKNOWN_CURRICULUM")
                if outcome_id not in outcomes or outcomes.get(outcome_id) != curriculum_id:
                    qerrors.append("INVALID_OUTCOME_LINK")
            if not sources[qid]:
                qerrors.append("MISSING_SOURCE")
            elif any(str(source.get("FILE_ID")) not in file_ids for source in sources[qid]):
                qerrors.append("MISSING_ASSET")
            raw_row = raw_by_id.get(qid, {})
            formula_fields = {}
            for field in ("DIFFICULTY", "COMPETENCY", "USE_PURPOSE", "SHUFFLE_ALLOWED",
                          "FAMILY_ID", "PROCESS_STATUS", "CLASSIFICATION_BASIS"):
                raw_value = raw_row.get(field)
                if isinstance(raw_value, str) and raw_value.startswith("="):
                    formula_fields[field] = raw_value
                    if row.get(field) is None:
                        qerrors.append(f"MISSING_CACHED_{field}")
            if qerrors:
                errors.append({"code": "INVALID_QUESTION", "question_id": qid, "issues": qerrors})
                continue
            normalized_options = [
                {"label": item.get("OPTION_LABEL"), "text": item.get("OPTION_TEXT"),
                 "is_correct": _boolean(item.get("IS_CORRECT")), "order": int(item.get("ORDER_NO"))}
                for item in sorted(options[qid], key=lambda item: int(item.get("ORDER_NO") or 0))
            ]
            normalized_statements = [
                {"label": item.get("STATEMENT_LABEL"), "text": item.get("STATEMENT_TEXT"),
                 "truth_value": _boolean(item.get("TRUTH_VALUE")), "order": int(item.get("ORDER_NO")),
                 "cognitive_level": item.get("COGNITIVE_LEVEL"), "difficulty": item.get("DIFFICULTY")}
                for item in sorted(statements[qid], key=lambda item: int(item.get("ORDER_NO") or 0))
            ]
            mapping = mappings[qid]
            canonical = {
                "question_id": qid, "question_code": row.get("QUESTION_CODE"), "question_type": qtype,
                "cognitive_level": row.get("COGNITIVE_LEVEL"), "stem_text": row.get("STEM_TEXT"),
                "answer_key": row.get("ANSWER_KEY"), "source_version": str(row.get("VERSION")),
                "options": normalized_options, "statements": normalized_statements,
                "curriculum_id": mapping.get("CURRICULUM_ID"), "outcome_id": mapping.get("OUTCOME_ID"),
                "difficulty": difficulty, "competency": row.get("COMPETENCY"),
                "estimated_time_seconds": row.get("ESTIMATED_TIME_SEC"),
                "process_status": row.get("PROCESS_STATUS"), "use_purpose": row.get("USE_PURPOSE"),
                "shuffle_allowed": shuffle_allowed, "family_id": row.get("FAMILY_ID"),
            }
            content_hash = hashlib.sha256(
                json.dumps(canonical, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()
            ).hexdigest()
            canonical.update({
                "content_hash": content_hash, "row": row, "mapping": mapping,
                "sources": sources[qid], "formula_fields": formula_fields,
            })
            result.append(canonical)
        return result
